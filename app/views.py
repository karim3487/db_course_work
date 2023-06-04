import locale
from datetime import datetime

from django.db import models
from django.db.models import Sum, F, Value, Subquery, OuterRef
from django.db.models.functions import Concat, Left
from django.http import HttpResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.utils import formats
from django.views.generic import CreateView, ListView, DeleteView, UpdateView
from openpyxl import Workbook

from app.forms import (
    DoctorCreationForm,
    PatientCreationForm,
    InsuranceCompanyCreationForm,
    AppointmentCreationForm,
    BillCreationForm,
    PaymentCreationForm,
)
from app.models import Doctor, Patient, InsuranceCompany, Bill, Payment, Appointment
from common.views import TitleMixin


def index(request):
    return HttpResponse("Hello, world. You're at the polls index.")


# DOCTOR:---------------------------------------------------------------------------------------------------------------
class DoctorListView(TitleMixin, ListView):
    model = Doctor
    template_name = "doctor/list.html"
    title = "Врачи"
    header = "Таблица с врачами"

    def get_queryset(self):
        queryset = super(DoctorListView, self).get_queryset()
        return queryset.all()

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(DoctorListView, self).get_context_data()
        return context


class DoctorCreateView(TitleMixin, CreateView):
    model = Doctor
    form_class = DoctorCreationForm
    template_name = "doctor/create.html"
    success_url = reverse_lazy("hospital:doctors")
    title = "Добавление врача"


class DoctorUpdateView(TitleMixin, UpdateView):
    model = Doctor
    form_class = DoctorCreationForm
    title = "Изменение врача"
    template_name = "doctor/update.html"
    success_url = reverse_lazy("hospital:doctors")


class DoctorDeleteView(TitleMixin, DeleteView):
    model = Doctor
    success_url = reverse_lazy("hospital:doctors")
    template_name = "confirm_delete.html"
    title = "Удаление врача"


def export_doctors(request):
    queryset = Doctor.objects.all()

    workbook = Workbook()
    sheet = workbook.active

    headers = ["ФИО", "Специальность", "Номер телефона", "E-mail", "Адрес"]
    sheet.append(headers)

    for item in queryset:
        row = [
            item.full_name,
            item.speciality,
            item.phone_number,
            item.email,
            item.address,
        ]
        sheet.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=doctors.xlsx"

    workbook.save(response)

    return response


# PATIENT:--------------------------------------------------------------------------------------------------------------
class PatientListView(TitleMixin, ListView):
    model = Patient
    template_name = "patient/list.html"
    title = "Пациенты"
    header = "Таблица с пациентами"

    def get_queryset(self):
        queryset = super(PatientListView, self).get_queryset()
        return queryset.all()

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(PatientListView, self).get_context_data()
        return context


class PatientCreateView(TitleMixin, CreateView):
    model = Patient
    form_class = PatientCreationForm
    template_name = "patient/create.html"
    success_url = reverse_lazy("hospital:patients")
    title = "Добавление пациента"


class PatientUpdateView(TitleMixin, UpdateView):
    model = Patient
    form_class = PatientCreationForm
    title = "Изменение пациента"
    template_name = "patient/update.html"
    success_url = reverse_lazy("hospital:patients")


class PatientDeleteView(TitleMixin, DeleteView):
    model = Patient
    success_url = reverse_lazy("hospital:patients")
    template_name = "confirm_delete.html"
    title = "Удаление пациента"


def export_patients(request):
    queryset = Patient.objects.all()

    workbook = Workbook()
    sheet = workbook.active

    headers = ["ФИО", "Номер телефона", "Адрес", "Страховая компания"]
    sheet.append(headers)

    for item in queryset:
        row = [
            item.full_name,
            item.phone_number,
            item.address,
            item.insurance_company.name if item.insurance_company else "Не застрахован",
        ]
        sheet.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=patients.xlsx"

    workbook.save(response)

    return response


# INS_COMPANIES---------------------------------------------------------------------------------------------------------
class InsuranceCompanyListView(TitleMixin, ListView):
    model = InsuranceCompany
    template_name = "ins_company/list.html"
    title = "Страховые компании"
    header = "Таблица со страховыми компаниями"

    def get_queryset(self):
        queryset = super(InsuranceCompanyListView, self).get_queryset()
        return queryset.all()

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(InsuranceCompanyListView, self).get_context_data()
        return context


class InsuranceCompanyCreateView(TitleMixin, CreateView):
    model = InsuranceCompany
    form_class = InsuranceCompanyCreationForm
    template_name = "ins_company/create.html"
    success_url = reverse_lazy("hospital:ins_companies")
    title = "Добавление страховой компании"


class InsuranceCompanyUpdateView(TitleMixin, UpdateView):
    model = InsuranceCompany
    form_class = InsuranceCompanyCreationForm
    title = "Изменение страховой компании"
    template_name = "ins_company/update.html"
    success_url = reverse_lazy("hospital:ins_companies")


class InsuranceCompanyDeleteView(TitleMixin, DeleteView):
    model = InsuranceCompany
    success_url = reverse_lazy("hospital:ins_companies")
    template_name = "confirm_delete.html"
    title = "Удаление страховой компании"


def export_ins_companies(request):
    queryset = InsuranceCompany.objects.all()

    workbook = Workbook()
    sheet = workbook.active

    headers = ["Название", "Номер телефона", "Адрес"]
    sheet.append(headers)

    for item in queryset:
        row = [
            item.name,
            item.phone_number,
            item.address,
        ]
        sheet.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=companies.xlsx"

    workbook.save(response)

    return response


# APPOINTMENT:----------------------------------------------------------------------------------------------------------
class AppointmentListView(TitleMixin, ListView):
    model = Appointment
    template_name = "appointment/list.html"
    title = "Приемы"
    header = "Таблица с приемами"

    def get_queryset(self):
        queryset = super(AppointmentListView, self).get_queryset()
        return queryset.all()

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(AppointmentListView, self).get_context_data()
        return context


class AppointmentCreateView(TitleMixin, CreateView):
    model = Appointment
    form_class = AppointmentCreationForm
    template_name = "appointment/create.html"
    success_url = reverse_lazy("hospital:appointments")
    title = "Добавление приема"


class AppointmentUpdateView(TitleMixin, UpdateView):
    model = Appointment
    form_class = AppointmentCreationForm
    title = "Изменение приема"
    template_name = "appointment/update.html"
    success_url = reverse_lazy("hospital:appointments")


class AppointmentDeleteView(TitleMixin, DeleteView):
    model = Appointment
    success_url = reverse_lazy("hospital:appointments")
    template_name = "confirm_delete.html"
    title = "Удаление приема"


def export_appointments(request):
    locale.setlocale(locale.LC_TIME, "ru_RU")
    queryset = Appointment.objects.all()

    workbook = Workbook()
    sheet = workbook.active

    headers = ["Доктор", "Пациент", "Дата и время"]
    sheet.append(headers)

    for item in queryset:
        row = [
            str(item.doctor),
            str(item.patient),
            str(formats.date_format(item.datetime, format="d.m.Y, H:i")),
        ]
        sheet.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=appointments.xlsx"

    workbook.save(response)

    return response


# BILL:-----------------------------------------------------------------------------------------------------------------
class BillListView(TitleMixin, ListView):
    model = Bill
    template_name = "bill/list.html"
    title = "Счета"
    header = "Таблица со счетами"

    def get_queryset(self):
        queryset = super(BillListView, self).get_queryset()
        return queryset.all()

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(BillListView, self).get_context_data()
        return context


class BillCreateView(TitleMixin, CreateView):
    model = Bill
    form_class = BillCreationForm
    template_name = "bill/create.html"
    success_url = reverse_lazy("hospital:bills")
    title = "Добавление счета"


class BillUpdateView(TitleMixin, UpdateView):
    model = Bill
    form_class = BillCreationForm
    title = "Изменение счета"
    template_name = "bill/update.html"
    success_url = reverse_lazy("hospital:bills")


class BillDeleteView(TitleMixin, DeleteView):
    model = Bill
    success_url = reverse_lazy("hospital:bills")
    template_name = "confirm_delete.html"
    title = "Удаление счета"


def export_bills(request):
    queryset = Bill.objects.all()

    workbook = Workbook()
    sheet = workbook.active

    headers = ["Прием", "Застрахован ли пациент", "Дата отправки счета", "Сумма"]
    sheet.append(headers)

    for item in queryset:
        row = [
            str(item.appointment),
            "Да" if item.is_amount_insured else "Нет",
            str(formats.date_format(item.date_sent, format="d.m.Y")),
            item.amount,
        ]
        sheet.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=bills.xlsx"

    workbook.save(response)

    return response


# PAYMENT:-----------------------------------------------------------------------------------------------------------------
class PaymentListView(TitleMixin, ListView):
    model = Payment
    template_name = "payment/list.html"
    title = "Платежи"
    header = "Таблица с платежами"

    def get_queryset(self):
        queryset = super(PaymentListView, self).get_queryset()
        return queryset.all()

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(PaymentListView, self).get_context_data()
        return context


class PaymentCreateView(TitleMixin, CreateView):
    model = Payment
    form_class = PaymentCreationForm
    template_name = "payment/create.html"
    title = "Добавление платежа"

    def post(self, request, *args, **kwargs):
        data = request.POST
        bill = Bill.objects.get(id=data["bill"])
        patient = bill.appointment.patient
        ins_company = patient.insurance_company

        date_str = "04.06.2023"
        date_obj = datetime.strptime(date_str, "%d.%m.%Y")
        formatted_date_str = date_obj.strftime("%Y-%m-%d")
        payment = Payment(
            patient=patient,
            insurance_company=ins_company,
            bill=bill,
            amount=data["amount"],
            date=formatted_date_str,
        )
        payment.save()
        return redirect("hospital:payments")


class PaymentUpdateView(TitleMixin, UpdateView):
    model = Payment
    form_class = PaymentCreationForm
    title = "Изменение платежа"
    template_name = "payment/update.html"
    success_url = reverse_lazy("hospital:payments")


class PaymentDeleteView(TitleMixin, DeleteView):
    model = Payment
    success_url = reverse_lazy("hospital:payments")
    template_name = "confirm_delete.html"
    title = "Удаление платежа"


def export_payments(request):
    queryset = Payment.objects.all()

    workbook = Workbook()
    sheet = workbook.active

    headers = ["Пациент", "Страховая компания", "Счет", "Сумма"]
    sheet.append(headers)

    for item in queryset:
        row = [
            str(item.patient),
            "Не застрахован" if item.insurance_company else str(item.insurance_company),
            item.bill,
            item.amount,
        ]
        sheet.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=payment.xlsx"

    workbook.save(response)

    return response


# BILL_PAYMENTS:--------------------------------------------------------------------------------------------------------
class BillPaymentListView(TitleMixin, ListView):
    template_name = "bill_payments/list.html"
    context_object_name = "bill_payments"
    title = "Счета и платежи"
    header = "Таблица с счетами и платежами"

    def get_queryset(self):
        total_payment_subquery = (
            Payment.objects.filter(bill_id=OuterRef("id"))
            .values("bill_id")
            .annotate(total_payment=Sum("amount"))
            .values("total_payment")
        )

        queryset = (
            Bill.objects.select_related("appointment__doctor")
            .annotate(
                total_payment=Subquery(
                    total_payment_subquery, output_field=models.DecimalField()
                )
            )
            .annotate(
                balance=F("amount") - F("total_payment"),
                doctor=Concat(
                    F("appointment__doctor__first_name"),
                    Value(" "),
                    Left(F("appointment__doctor__last_name"), 1),
                    Value(". "),
                    Left(F("appointment__doctor__surname"), 1),
                    Value(". — "),
                    F("appointment__doctor__speciality"),
                ),
                patient=Concat(
                    F("appointment__patient__first_name"),
                    Value(" "),
                    Left(F("appointment__patient__last_name"), 1),
                    Value(". "),
                    Left(F("appointment__patient__surname"), 1),
                    Value("."),
                ),
            )
        )
        # queryset = Bill.objects.annotate(total_payment=Sum("payment__amount")).annotate(
        #     balance=F("amount") - F("total_payment")
        # ).annotate(doctor=Concat(F("appointment__doctor__first_name"), Value(" "), F("appointment__doctor__last_name"),
        #                          Value(" "), F("appointment__doctor__surname")))
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        total_payment_sum = Payment.objects.aggregate(total=Sum("amount"))["total"]
        context["total_payment_sum"] = total_payment_sum
        return context


def export_bill_payments(request):
    queryset = Bill.objects.annotate(total_payment=Sum("payment__amount")).annotate(
        balance=F("amount") - F("total_payment")
    )
    total_payment_sum = Payment.objects.aggregate(total=Sum("amount"))["total"]

    workbook = Workbook()
    sheet = workbook.active

    headers = [
        "Доктор",
        "Пациент",
        "Сумма за прием",
        "Дата отправки счета",
        "Общая сумма платежей",
        "Остаток",
    ]
    sheet.append(headers)

    for item in queryset:
        row = [
            str(item.appointment.doctor),
            str(item.appointment.patient),
            item.amount,
            item.date_sent,
            item.total_payment,
            item.balance if item.balance > 0 else "Оплачено",
        ]
        sheet.append(row)

    sheet.append(["Всего заработано", total_payment_sum])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=bill_payments.xlsx"

    workbook.save(response)

    return response
